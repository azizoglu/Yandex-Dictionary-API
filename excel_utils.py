from openpyxl import load_workbook
import pandas as pd
import json
import yandex_api_utils

with open('config/config.json') as f:
    config = json.load(f)

API_KEY = config["API_KEY"]
BASE_URL = config["BASE_URL"]
PARAM_LANG = config["PARAM_LANG"]
PARAM_TEXT = config["PARAM_TEXT"]

# Insert The Synonyms of Text in The Third Column of Excel
def insertSynonyms (file_name, sheet_name) :
    wb = load_workbook(filename=file_name, read_only=False)
    ws = wb[sheet_name]

    i = 0
    for row in ws.rows:
       i = i + 1
       value=row[0].value
       if i != 1:
            REQUEST_TEXT = value
            REQUEST_URL = BASE_URL + API_KEY + PARAM_LANG + PARAM_TEXT + REQUEST_TEXT   
            content = yandex_api_utils.getResultOfRequest(REQUEST_URL)
            synonym = yandex_api_utils.getSynonyms(content)
            if synonym == '':
                print(value + ' synonym is not exists')
                ws.cell(row=i, column=3).value = value + ' synonym is not exists'
            else :
                print(synonym)
                ws.cell(row=i, column=3).value = synonym


    wb.save(file_name)

# Insert The Meanings of Text in The Third Column of Excel
def insertMeanings (file_name, sheet_name) :
    wb = load_workbook(filename=file_name, read_only=False)
    ws = wb[sheet_name]

    i = 0
    for row in ws.rows:
       i = i + 1
       value=row[0].value
       if i != 1:
            REQUEST_TEXT = value
            REQUEST_URL = BASE_URL + API_KEY + PARAM_LANG + PARAM_TEXT + REQUEST_TEXT   
            content = yandex_api_utils.getResultOfRequest(REQUEST_URL)
            mean = yandex_api_utils.getMeans(content)
            if mean == '':
                print(value + ' mean is not exists')
                ws.cell(row=i, column=2).value = value + ' mean is not exists'
            else :
                print(mean)
                ws.cell(row=i, column=2).value = mean


    wb.save(file_name) 